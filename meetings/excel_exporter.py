import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


class MeetingExcelExporter:
    """
    Export meeting data to Excel with formatting
    """
    
    def __init__(self, meeting):
        self.meeting = meeting
        self.wb = openpyxl.Workbook()
        
    def generate_excel(self):
        """Generate Excel workbook"""
        # Remove default sheet
        self.wb.remove(self.wb.active)
        
        # Add sheets
        self._add_overview_sheet()
        self._add_speaker_metrics_sheet()
        self._add_messages_sheet()
        
        # Save to buffer
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    def _add_overview_sheet(self):
        """Add overview sheet"""
        ws = self.wb.create_sheet("Overview")
        
        # Header style
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Title
        ws['A1'] = "TeamPulse AI - Meeting Analysis"
        ws['A1'].font = Font(bold=True, size=16, color="4F46E5")
        ws.merge_cells('A1:B1')
        
        # Meeting info
        ws['A3'] = "Meeting Information"
        ws['A3'].font = header_font
        ws['A3'].fill = header_fill
        ws.merge_cells('A3:B3')
        
        data = [
            ['Title', self.meeting.title],
            ['Date', self.meeting.date.strftime('%B %d, %Y at %I:%M %p')],
            ['Duration', f"{self.meeting.duration_minutes} minutes" if self.meeting.duration_minutes else 'N/A'],
            ['Total Messages', self.meeting.total_messages],
            ['Total Words', self.meeting.total_words],
            ['Average Sentiment', f"{self.meeting.avg_sentiment:.4f}" if self.meeting.avg_sentiment else 'N/A'],
            ['Sentiment Label', self.meeting.sentiment_label.title()],
            ['Participation Balance', f"{self.meeting.participation_balance:.4f}" if self.meeting.participation_balance else 'N/A'],
        ]
        
        row = 4
        for label, value in data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
    
    def _add_speaker_metrics_sheet(self):
        """Add speaker metrics sheet"""
        ws = self.wb.create_sheet("Speaker Metrics")
        
        # Header style
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Headers
        headers = ['Speaker', 'Messages', 'Words', 'Participation %', 'Avg Words/Msg', 
                   'Engagement Score', 'Avg Sentiment', 'Questions', 'Positive', 'Neutral', 'Negative']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        metrics = self.meeting.metrics.all().order_by('-participation_percentage')
        
        for row, metric in enumerate(metrics, 2):
            ws.cell(row=row, column=1, value=metric.speaker.name)
            ws.cell(row=row, column=2, value=metric.total_messages)
            ws.cell(row=row, column=3, value=metric.total_words)
            ws.cell(row=row, column=4, value=f"{metric.participation_percentage:.2f}")
            ws.cell(row=row, column=5, value=f"{metric.avg_words_per_message:.2f}")
            ws.cell(row=row, column=6, value=f"{metric.engagement_score:.2f}")
            ws.cell(row=row, column=7, value=f"{metric.avg_sentiment:.4f}" if metric.avg_sentiment else 'N/A')
            ws.cell(row=row, column=8, value=metric.question_count)
            ws.cell(row=row, column=9, value=metric.positive_count)
            ws.cell(row=row, column=10, value=metric.neutral_count)
            ws.cell(row=row, column=11, value=metric.negative_count)
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _add_messages_sheet(self):
        """Add individual messages sheet"""
        ws = self.wb.create_sheet("Messages")
        
        # Header style
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Headers
        headers = ['#', 'Speaker', 'Content', 'Word Count', 'Sentiment', 'Is Question']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        messages = self.meeting.messages.all().order_by('sequence_order')
        
        for row, msg in enumerate(messages, 2):
            ws.cell(row=row, column=1, value=msg.sequence_order)
            ws.cell(row=row, column=2, value=msg.speaker.name)
            ws.cell(row=row, column=3, value=msg.content)
            ws.cell(row=row, column=4, value=msg.word_count)
            ws.cell(row=row, column=5, value=f"{msg.sentiment_score:.4f}" if msg.sentiment_score else 'N/A')
            ws.cell(row=row, column=6, value='Yes' if msg.is_question else 'No')
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12


class TeamHealthExcelExporter:
    """
    Export team health data across multiple meetings
    """
    
    def __init__(self, meetings):
        self.meetings = meetings
        self.wb = openpyxl.Workbook()
    
    def generate_excel(self):
        """Generate Excel workbook"""
        self.wb.remove(self.wb.active)
        
        self._add_summary_sheet()
        self._add_meetings_list_sheet()
        self._add_trends_sheet()
        
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    def _add_summary_sheet(self):
        """Add summary sheet"""
        ws = self.wb.create_sheet("Summary")
        
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        ws['A1'] = "Team Health Report"
        ws['A1'].font = Font(bold=True, size=16, color="4F46E5")
        ws.merge_cells('A1:B1')
        
        ws['A3'] = "Analysis Period"
        ws['A3'].font = header_font
        ws['A3'].fill = header_fill
        ws.merge_cells('A3:B3')
        
        ws['A4'] = "From"
        ws['B4'] = self.meetings[-1].date.strftime('%B %d, %Y')
        ws['A5'] = "To"
        ws['B5'] = self.meetings[0].date.strftime('%B %d, %Y')
        ws['A6'] = "Total Meetings"
        ws['B6'] = len(self.meetings)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _add_meetings_list_sheet(self):
        """Add meetings list"""
        ws = self.wb.create_sheet("Meetings")
        
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        headers = ['Date', 'Title', 'Messages', 'Words', 'Sentiment', 'Balance']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row, meeting in enumerate(self.meetings, 2):
            ws.cell(row=row, column=1, value=meeting.date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=meeting.title)
            ws.cell(row=row, column=3, value=meeting.total_messages)
            ws.cell(row=row, column=4, value=meeting.total_words)
            ws.cell(row=row, column=5, value=f"{meeting.avg_sentiment:.4f}" if meeting.avg_sentiment else 'N/A')
            ws.cell(row=row, column=6, value=f"{meeting.participation_balance:.4f}" if meeting.participation_balance else 'N/A')
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _add_trends_sheet(self):
        """Add trends data"""
        from .trend_analyzer import TrendAnalyzer
        
        ws = self.wb.create_sheet("Trends")
        
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        trend_analyzer = TrendAnalyzer(self.meetings)
        trends = trend_analyzer.get_comprehensive_analysis()
        
        ws['A1'] = "Trend Analysis"
        ws['A1'].font = Font(bold=True, size=14, color="4F46E5")
        
        row = 3
        for metric_name, trend_data in trends.items():
            if metric_name == 'meeting_count':
                continue
            
            ws[f'A{row}'] = metric_name.replace('_', ' ').title()
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            row += 1
            
            ws[f'A{row}'] = "Trend"
            ws[f'B{row}'] = trend_data.get('trend', 'N/A')
            row += 1
            
            ws[f'A{row}'] = "Change %"
            ws[f'B{row}'] = f"{trend_data.get('change_percentage', 0):.2f}%"
            row += 1
            
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        